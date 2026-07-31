"""'Veli İletişim Bilgileri' Excel'ini okuma: başlık tespiti, esnek (fuzzy)
sütun eşleme, satır ayrıştırma + normalize.

Beklenen sütunlar (örnek dosyadan):
  | Sınıf | TCKN | Numa | Adı Soyadı | Veli Kim
  | Anne Adı SOYADI | AnneTel | Baba Adı SOYADI | BabaTel |

OYS `apps/core/excel_veli.py` dosyasından AYNEN alındı (tasarım §4.4); yalnız
import yolu uyarlandı. DB eşleştirme/yazma `services/imports.py`'dadır (saf modül).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from apps.okul import normalize

# Türkçe karakter → ASCII (başlık eşlemesi için; küçük harfe indirgenir).
_TR_MAP = str.maketrans(
    {
        "ş": "s",
        "Ş": "s",
        "ı": "i",
        "I": "i",
        "İ": "i",
        "ğ": "g",
        "Ğ": "g",
        "ü": "u",
        "Ü": "u",
        "ö": "o",
        "Ö": "o",
        "ç": "c",
        "Ç": "c",
    }
)
_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def normalize_header(text: Any) -> str:
    """Başlık hücresini eşleme için normalleştirir ('Anne Adı SOYADI' → 'anne adi soyadi')."""
    if text is None:
        return ""
    s = str(text).translate(_TR_MAP).lower()
    s = _NON_ALNUM.sub(" ", s)
    return s.strip()


# Mantıksal alan → normalize edilmiş başlık anahtar kelimeleri.
# SIRA KRİTİK: telefon alanları isim alanlarından ÖNCE ('annetel' → 'anne'
# yanlış eşleşmesini önlemek için); veli isimleri öğrenci isimlerinden önce.
COLUMN_SYNONYMS: dict[str, list[str]] = {
    "class": ["sinif sube", "sinif/sube", "sinif", "sube"],
    "tckn": ["tckn", "tc kimlik", "t c kimlik", "kimlik no", "tc no"],
    "number": ["okul no", "okul numarasi", "ogrenci no", "ogrenci numarasi", "numara", "numa"],
    "guardian": ["veli kim", "sorumlu veli", "veli"],
    "mother_phone": ["annetel", "anne tel", "anne telefon", "anne gsm", "anne cep", "anne no"],
    "father_phone": ["babatel", "baba tel", "baba telefon", "baba gsm", "baba cep", "baba no"],
    "mother_name": ["anne adi soyadi", "anne adi", "anne ad soyad", "anne"],
    "father_name": ["baba adi soyadi", "baba adi", "baba ad soyad", "baba"],
    "student_name": [
        "ogrenci adi soyadi",
        "adi soyadi",
        "ad soyad",
        "adsoyad",
        "ad ve soyad",
        "isim",
    ],
    "student_first": ["ogrenci adi"],
    "student_last": ["ogrenci soyadi"],
    # Standart OYS şablonu ek sütunları (Tur 531, ADR-0034) — e-Okul ihracında
    # bulunmaz, opsiyoneldir; CRITICAL_FIELDS değişmediği için e-Okul akışı aynen.
    # Listenin SONUNDA: telefon/isim öncelik sırası bozulmaz.
    "birth_date": ["dogum tarihi"],
    "gender": ["cinsiyet"],
}

# Bu sütunlar olmadan içe aktarma yapılamaz (ParserError).
CRITICAL_FIELDS = ("class",)


class ParserError(Exception):
    """Excel ayrıştırılamadığında fırlatılır (kritik sütun eksik vb.)."""


@dataclass
class ColumnMapping:
    """Tespit edilen başlık satırı ve alan → kolon indeksi eşlemesi."""

    header_row: int
    fields: dict[str, int] = field(default_factory=dict)
    matched_headers: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    @property
    def missing_critical(self) -> list[str]:
        missing = [f for f in CRITICAL_FIELDS if f not in self.fields]
        if "tckn" not in self.fields and "number" not in self.fields:
            missing.append("number veya tckn")
        has_combined_name = "student_name" in self.fields
        has_split_name = "student_first" in self.fields and "student_last" in self.fields
        if not has_combined_name and not has_split_name:
            missing.append("student_name veya student_first+student_last")
        return missing

    @property
    def is_usable(self) -> bool:
        return not self.missing_critical


@dataclass
class ParsedParent:
    """Tek bir velinin (anne/baba) çözümlenmiş hâli."""

    raw_name: str = ""
    first_name: str = ""
    last_name: str = ""
    raw_phone: str = ""
    phone: str | None = None  # normalize edilmiş ('05XXXXXXXXX') veya None

    @property
    def has_any_data(self) -> bool:
        """Bu veli için dosyada herhangi bir bilgi (ad veya telefon) var mı?"""
        return bool(self.raw_name.strip()) or bool(self.raw_phone.strip())


@dataclass
class ParsedRow:
    """Bir veri satırının normalize edilmiş hâli (DB eşleştirmesi öncesi)."""

    row_number: int  # 1-tabanlı (Excel satır no'su)
    raw_class: str = ""
    class_level: int | None = None
    class_section: str = ""
    raw_tckn: str = ""
    tckn: str | None = None  # normalize + checksum geçerli ise; aksi None
    student_number: str = ""
    raw_student_name: str = ""
    student_first: str = ""
    student_last: str = ""
    guardian: str = ""  # ANNE / BABA / DIGER / ""
    mother: ParsedParent = field(default_factory=ParsedParent)
    father: ParsedParent = field(default_factory=ParsedParent)
    # Standart şablon ek alanları (Tur 531, ADR-0034) — e-Okul dosyasında boş kalır.
    raw_birth_date: str = ""
    birth_date: date | None = None  # normalize_excel_date çözümü; çözülemezse None
    raw_gender: str = ""
    gender: str = ""  # normalize_gender: 'E' / 'K' / ''


def _match_field_for_header(norm: str) -> str | None:
    """Normalize başlığı bir mantıksal alana eşler (ilk eşleşen kazanır)."""
    for fieldname, keywords in COLUMN_SYNONYMS.items():
        for kw in keywords:
            if kw == norm or kw in norm:
                return fieldname
    return None


def _map_header_row(cells: list[Any]) -> tuple[dict[str, int], dict[str, str], list[str]]:
    fields: dict[str, int] = {}
    matched: dict[str, str] = {}
    warnings: list[str] = []
    for idx, cell in enumerate(cells):
        norm = normalize_header(cell)
        if not norm:
            continue
        fieldname = _match_field_for_header(norm)
        if fieldname is None:
            continue
        if fieldname in fields:
            warnings.append(
                f"'{cell}' sütunu '{fieldname}' için yinelenen eşleşme; "
                f"ilk eşleşen '{matched[fieldname]}' kullanıldı."
            )
            continue
        fields[fieldname] = idx
        matched[fieldname] = str(cell).strip()
    return fields, matched, warnings


def detect_columns(rows: list[list[Any]], scan_limit: int = 10) -> ColumnMapping:
    """İlk satırlar içinde en iyi başlık satırını bulur ve sütunları eşler."""
    best: ColumnMapping | None = None
    for r in range(min(scan_limit, len(rows))):
        cells = rows[r]
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        fields, matched, warnings = _map_header_row(cells)
        if not fields:
            continue
        candidate = ColumnMapping(
            header_row=r, fields=fields, matched_headers=matched, warnings=warnings
        )
        if best is None or (candidate.is_usable, len(candidate.fields)) > (
            best.is_usable,
            len(best.fields),
        ):
            best = candidate
        if best.is_usable and best.header_row == r:
            break
    return best or ColumnMapping(header_row=0)


def _cell(cells: list[Any], idx: int | None) -> Any:
    if idx is None or idx >= len(cells):
        return None
    return cells[idx]


def _str(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    # Excel sayıyı float metnine çevirebilir ('2612.0' → '2612').
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


_GUARDIAN_MAP = {
    "ANNE": "ANNE",
    "BABA": "BABA",
    "MOTHER": "ANNE",
    "FATHER": "BABA",
}


def _normalize_guardian(value: Any) -> str:
    """'ANNE'/'BABA' → aynısı; 'DİĞER'/tanınmayan → 'DIGER'; boş → ''."""
    s = _str(value)
    if not s:
        return ""
    key = normalize_header(s).upper().replace(" ", "")
    if key in _GUARDIAN_MAP:
        return _GUARDIAN_MAP[key]
    return "DIGER"


def _parse_parent(
    cells: list[Any], f: dict[str, int], name_key: str, phone_key: str
) -> ParsedParent:
    raw_name = _str(_cell(cells, f.get(name_key)))
    raw_phone = _str(_cell(cells, f.get(phone_key)))
    first, last = normalize.split_full_name(raw_name)
    return ParsedParent(
        raw_name=raw_name,
        first_name=first,
        last_name=last,
        raw_phone=raw_phone,
        phone=normalize.normalize_phone(raw_phone) if raw_phone else None,
    )


def parse_rows(rows: list[list[Any]], mapping: ColumnMapping) -> list[ParsedRow]:
    """Başlık satırından sonraki veri satırlarını çözümler (boş satırlar atlanır)."""
    f = mapping.fields
    parsed: list[ParsedRow] = []
    for r in range(mapping.header_row + 1, len(rows)):
        cells = rows[r]
        if all(c is None or str(c).strip() == "" for c in cells):
            continue

        raw_class = _str(_cell(cells, f.get("class")))
        class_parsed = normalize.normalize_class_section(raw_class)
        raw_tckn = _str(_cell(cells, f.get("tckn")))
        raw_name = _str(_cell(cells, f.get("student_name")))
        if raw_name:
            first, last = normalize.split_full_name(raw_name)
        else:
            first = _str(_cell(cells, f.get("student_first")))
            last = _str(_cell(cells, f.get("student_last")))
            raw_name = f"{first} {last}".strip()
        # Standart şablon ek alanları — hücre datetime da metin de olabilir;
        # normalize edici ikisini de kabul eder (raw metin rapor uyarıları için).
        birth_cell = _cell(cells, f.get("birth_date"))
        raw_gender = _str(_cell(cells, f.get("gender")))

        parsed.append(
            ParsedRow(
                row_number=r + 1,
                raw_class=raw_class,
                class_level=class_parsed[0] if class_parsed else None,
                class_section=class_parsed[1] if class_parsed else "",
                raw_tckn=raw_tckn,
                tckn=normalize.normalize_tckn(raw_tckn),
                student_number=_str(_cell(cells, f.get("number"))),
                raw_student_name=raw_name,
                student_first=first,
                student_last=last,
                guardian=_normalize_guardian(_cell(cells, f.get("guardian"))),
                mother=_parse_parent(cells, f, "mother_name", "mother_phone"),
                father=_parse_parent(cells, f, "father_name", "father_phone"),
                raw_birth_date=_str(birth_cell),
                birth_date=normalize.normalize_excel_date(birth_cell),
                raw_gender=raw_gender,
                gender=normalize.normalize_gender(raw_gender) if raw_gender else "",
            )
        )
    return parsed


def read_sheet(file_bytes: bytes) -> list[list[Any]]:
    """Excel baytlarını satır-listesine çevirir (etkin sayfa, salt-okunur)."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def parse_workbook(file_bytes: bytes) -> tuple[ColumnMapping, list[ParsedRow]]:
    """Baytlardan (mapping, satırlar) üretir; kritik sütun eksikse ParserError."""
    grid = read_sheet(file_bytes)
    mapping = detect_columns(grid)
    if not mapping.is_usable:
        eksik = ", ".join(mapping.missing_critical)
        raise ParserError(f"Zorunlu sütun(lar) bulunamadı: {eksik}.")
    return mapping, parse_rows(grid, mapping)
